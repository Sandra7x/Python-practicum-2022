from flask import Flask, render_template, request, send_file
import requests
from views import views
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from meteostat import Point, Daily
import pandas as pd
from openpyxl import load_workbook



app = Flask(__name__)
app.register_blueprint(views, url_prefix="/views")


@app.route('/weather', methods = ['GET', 'POST'])
def weather():
    if request.method == 'POST':
        city_name = request.form['city']
        APP_ID = "60aa068482d6ddc251ae5f53570ac5fb"
        URL = "https://api.openweathermap.org/data/2.5/weather"
        PARAMS = {'q': city_name, 'appid': APP_ID, 'units':'metric', 'lang': 'la'}
        response = requests.get(url=URL, params=PARAMS).json()

        temp = response['main']['temp']
        weather = response['weather'][0]['description']
        icon = response['weather'][0]['icon']
        lat = response['coord']['lat']
        lon = response['coord']['lon']
    
        return render_template('weather.html',temp=temp,weather=weather,icon=icon, city_name = city_name, lat = lat, lon = lon)
    else:
        return render_template('weather.html')


@app.route('/weather_history', methods = ['GET', 'POST'])
def weather_history():
    if request.method == 'POST':
        city_name = request.form['city']
        APP_ID = "60aa068482d6ddc251ae5f53570ac5fb"
        URL = "https://api.openweathermap.org/data/2.5/weather"
        PARAMS = {'q': city_name, 'appid': APP_ID, 'units':'metric', 'lang': 'la'}
        response = requests.get(url=URL, params=PARAMS).json()
        lat_coord = response['coord']['lat']
        lon_coord = response['coord']['lon']
        
        start = datetime(2022, 1, 1)
        end = datetime(2022, 12, 31)
        city_history = Point(lat_coord, lon_coord)
        data = Daily(city_history, start, end)
        data = data.fetch()
        data.plot(y=['tavg', 'tmin', 'tmax'], grid = True, label=['average', 'min', 'max'])
        
        df = pd.DataFrame(data, columns = ['tavg', 'tmin', 'tmax'])
        records_total = len(df.index)
        df.loc['Average temp:'] = pd.Series(round(df['tavg'].sum()/records_total, 2), index=['tavg'])
        average_temp = round(df['tavg'].sum()/records_total)
        df.to_excel("weather_data.xlsx", sheet_name=city_name)
        min_temp = df['tmin'].min()
        df2=df[df["tmin"] == min_temp]
        max_temp = df['tmax'].max()

        pp = PdfPages("chart.pdf")
        plt.title(f'Weather temperature in {city_name} 2022', pad=30)
        plt.ylabel(' Temperature °C ')
        plt.xlabel('Month')
        plt.legend(bbox_to_anchor=(1.07, 1.0), loc='upper left')
        plt.tight_layout()
        pp.savefig()
        pp.close()
        plt.show()

        book = load_workbook("weather_data.xlsx")
        sheet = book.active

        return render_template('weather_history.html', city_name=city_name, lat_coord=lat_coord, lon_coord=lon_coord, sheet = sheet, min_temp = min_temp, max_temp = max_temp, average_temp = average_temp) 
    
    else:
        book = load_workbook("weather_data.xlsx")
        sheet = book.active
        return render_template('weather_history.html', book = book, sheet = sheet)


def upload_form():
	return render_template('weather_data.html')


@app.route('/weather_data')
def weather_data():
    path = "weather_data.xlsx"
    return send_file(path, as_attachment=True)


@app.route('/weather_data_pdf')
def chart_pdf():
    path_pdf = "chart.pdf"
    return send_file(path_pdf, as_attachment=True)



if __name__ == "__main__":
    app.run(debug=True)

